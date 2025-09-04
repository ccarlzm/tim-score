from __future__ import annotations

import csv
import re
import unicodedata
from datetime import datetime, date
from pathlib import Path
from typing import Optional, Tuple

from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from django.db import transaction
from django.utils import timezone

from openpyxl import load_workbook

from compcore.apps.accounts.models import Profile
from compcore.apps.events.models import Event, Division
from compcore.apps.registration.models import Team, AthleteEntry


# ======================
# Utilidades de nombres
# ======================

def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))

def _to_username_slug(full_name: str) -> str:
    """
    Genera un username base:
    - minúsculas
    - sin acentos
    - solo [a-z0-9_]
    - espacios/puntuación -> guiones bajos
    """
    s = full_name.strip().lower()
    s = _strip_accents(s)
    s = re.sub(r"[^a-z0-9]+", "_", s)         # cualquier no alfanum a _
    s = s.strip("_")
    return s or "user"

def _split_full_name(full_name: str) -> tuple[str, str]:
    """
    Separa un 'Nombre Apellido ...' en (first_name, last_name).
    - Si hay una sola palabra -> first_name=full, last_name=""
    - Si hay 2+ -> first_name=primera, last_name=resto
    """
    parts = full_name.strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])

def _parse_date(value) -> Optional[date]:
    """
    Acepta:
    - date ya convertido por openpyxl
    - string 'YYYY-MM-DD' o 'DD/MM/YYYY' o 'MM/DD/YYYY' (heurísticas)
    """
    if value in (None, "", "nan", "NaT"):
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    try:
        nums = re.findall(r"\d+", s)
        if len(nums) == 3:
            y, m, d = None, None, None
            if len(nums[0]) == 4:
                y, m, d = int(nums[0]), int(nums[1]), int(nums[2])
            else:
                d, m, y = int(nums[0]), int(nums[1]), int(nums[2])
            return date(y, m, d)
    except Exception:
        pass
    return None

# ======================
# Usuarios/Perfiles
# ======================

def _ensure_unique_username(base: str) -> str:
    candidate = base or "user"
    i = 1
    while User.objects.filter(username=candidate).exists():
        candidate = f"{base}{i}"
        i += 1
    return candidate

def _get_or_create_user_from_name_email(full_name: str, email: str | None) -> tuple[User, Optional[str], bool]:
    """
    Usa nombre completo + email para localizar/crear usuario.
    - Si existe por email → lo usa.
    - Si no existe → crea User con:
        username: slug del nombre (único)
        first_name/last_name: separados del nombre completo
        email: el provisto (si hay)
      Retorna (user, password_si_nuevo, created_bool)
    """
    full_name = (full_name or "").strip()
    email = (email or "").strip().lower()

    # Buscar por email si viene
    if email:
        try:
            u = User.objects.get(email=email)
            return u, None, False
        except User.DoesNotExist:
            pass

    # Crear uno nuevo
    base_username = _to_username_slug(full_name or (email.split("@")[0] if email else "user"))
    username = _ensure_unique_username(base_username)

    first_name, last_name = _split_full_name(full_name or username)
    new_password = User.objects.make_random_password(length=10)
    u = User(username=username, email=email or "")
    u.first_name = first_name
    u.last_name = last_name
    u.set_password(new_password)
    u.is_active = True
    u.save()

    Profile.objects.get_or_create(user=u)  # asegura Profile
    return u, new_password, True

def _update_profile(u: User, *, dob: Optional[date], id_doc: Optional[str]) -> None:
    prof, _ = Profile.objects.get_or_create(user=u)
    changed = False
    if dob and getattr(prof, "date_of_birth", None) != dob:
        prof.date_of_birth = dob
        changed = True
    if id_doc and getattr(prof, "id_document", "") != str(id_doc):
        prof.id_document = str(id_doc)
        changed = True
    if changed:
        prof.save(update_fields=["date_of_birth", "id_document"])

# ======================
# Divisiones/Eventos
# ======================

DIVISION_ALIASES = {
    # normaliza nombres frecuentes
    "masters": "Master",
    "master": "Master",
    "avanzado": "Avanzado",
    "novatos": "Novatos",
    "escalado": "Escalado",
    "funcional": "Funcional",
    "especial": "Especial",
}

def _strip_accents_lower(s: str) -> str:
    return _strip_accents((s or "").lower().strip())

def _normalize_division_name(name: str) -> str:
    key_low = _strip_accents_lower(name)
    return DIVISION_ALIASES.get(key_low, name.strip())

def _get_division(event: Event, division_name: str) -> Division:
    norm = _normalize_division_name(division_name)
    try:
        return Division.objects.get(event=event, name=norm)
    except Division.DoesNotExist:
        for d in Division.objects.filter(event=event):
            if _strip_accents_lower(d.name) == _strip_accents_lower(norm):
                return d
        raise CommandError(f"División '{division_name}' no existe en el evento '{event.slug}'.")

# ======================
# Importador
# ======================

COLUMNS = [
    "division_name",
    "team_name",
    "captain_username",
    "captain_Birth_date",
    "captain_ID",
    "captain_email",
    "member2_username",
    "member2_Birth_date",
    "member2_ID",
    "member2_email",
    "member3_username",
    "member3_Birth_date",
    "member3_ID",
    "member3_email",
    "member4_username",
    "member4_Birth_date",
    "member4_ID",
    "member4_email",
]

class Command(BaseCommand):
    help = "Importa equipos y miembros desde un .xlsx con nombres completos; crea usuarios y perfiles (DOB/ID)."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Ruta al archivo .xlsx con los equipos")
        parser.add_argument("--sheet", type=str, default=None, help="Nombre de la hoja (por defecto: primera)")
        parser.add_argument("--event-slug", required=True, help="Slug del evento destino (ej. force-games)")
        parser.add_argument("--dry-run", action="store_true", help="Simula sin escribir cambios")

    def handle(self, *args, **options):
        xlsx_path = Path(options["xlsx_path"])
        sheet_name = options.get("sheet")
        event_slug = options["event_slug"]
        dry_run = options.get("dry_run", False)

        if not xlsx_path.exists():
            raise CommandError(f"Archivo no encontrado: {xlsx_path}")

        try:
            event = Event.objects.get(slug=event_slug)
        except Event.DoesNotExist:
            raise CommandError(f"Evento '{event_slug}' no existe.")

        wb = load_workbook(filename=str(xlsx_path), data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

        # Validar cabecera
        header_cells = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        headers = [str(h).strip() if h is not None else "" for h in header_cells]

        for i, col in enumerate(COLUMNS):
            if i >= len(headers) or headers[i] != col:
                raise CommandError(
                    f"Cabecera inválida en columna {i+1}. Esperado '{col}', encontrado '{headers[i] if i < len(headers) else ''}'.\n"
                    f"Cabecera completa: {headers}"
                )

        # Preparar reporte
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = Path.cwd() / f"import_report_{timestamp}.csv"
        report_fp = None
        writer = None
        if not dry_run:
            report_fp = report_path.open("w", newline="", encoding="utf-8")
            writer = csv.writer(report_fp)
            writer.writerow([
                "row", "status", "division", "team_name", "created_users(user:pass)", "members_added", "warnings", "errors"
            ])

        total = ok = errs = warns = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            total += 1
            vals = [cell.value for cell in row]
            vals = [(str(v).strip() if v is not None else "") for v in vals]
            data = dict(zip(headers, vals))

            status = "OK"
            created_creds: list[str] = []
            members_added = 0
            warnings_list: list[str] = []
            errors_list: list[str] = []

            try:
                division = _get_division(event, data["division_name"])
                team_size = max(1, division.team_size)

                team_name = data["team_name"]
                if not team_name:
                    raise CommandError("team_name vacío.")

                # Capitán
                cap_fullname = data["captain_username"]
                cap_email = data["captain_email"]
                cap_dob = _parse_date(data["captain_Birth_date"])
                cap_id = data["captain_ID"]

                captain, cap_pw, cap_created = _get_or_create_user_from_name_email(cap_fullname, cap_email)
                if cap_created and cap_pw:
                    created_creds.append(f"{captain.username}:{cap_pw}")
                _update_profile(captain, dob=cap_dob, id_doc=cap_id)

                if not dry_run:
                    with transaction.atomic():
                        team, created = Team.objects.get_or_create(
                            event=event,
                            division=division,
                            name=team_name,
                            defaults={"captain": captain},
                        )
                        if not created and team.captain_id != captain.id:
                            team.captain = captain
                            team.save(update_fields=["captain"])

                        # Inscribir capitán
                        AthleteEntry.objects.get_or_create(
                            user=captain, event=event, division=division, team=team
                        )

                        remaining = max(0, team_size - team.member_count())

                        # Miembros 2..4
                        for n in (2, 3, 4):
                            full = data.get(f"member{n}_username", "")
                            email = data.get(f"member{n}_email", "")
                            dob = _parse_date(data.get(f"member{n}_Birth_date"))
                            id_doc = data.get(f"member{n}_ID", "")

                            if not full and not email:
                                continue
                            if remaining <= 0:
                                warnings_list.append("Equipo lleno; miembros adicionales ignorados.")
                                break

                            u, pw, created_u = _get_or_create_user_from_name_email(full, email)
                            if created_u and pw:
                                created_creds.append(f"{u.username}:{pw}")
                            _update_profile(u, dob=dob, id_doc=id_doc)

                            AthleteEntry.objects.get_or_create(
                                user=u, event=event, division=division, team=team
                            )
                            remaining -= 1
                            members_added += 1
                else:
                    # Simulación
                    simulated_total = 1  # capitán
                    for n in (2, 3, 4):
                        full = data.get(f"member{n}_username", "")
                        email = data.get(f"member{n}_email", "")
                        if full or email:
                            simulated_total += 1
                    members_added = max(0, min(simulated_total, team_size) - 1)

            except Exception as e:
                status = "ERROR"
                errors_list.append(str(e))
                errs += 1
            else:
                ok += 1
                warns += len(warnings_list)

            if writer:
                writer.writerow([
                    idx,
                    status,
                    data["division_name"],
                    data["team_name"],
                    ";".join(created_creds),
                    members_added,
                    "; ".join(warnings_list),
                    "; ".join(errors_list),
                ])

        if report_fp:
            report_fp.close()

        self.stdout.write(self.style.SUCCESS(f"Filas procesadas: {total}"))
        self.stdout.write(self.style.SUCCESS(f"OK: {ok}  ·  ERRORES: {errs}  ·  WARNINGS: {warns}"))
        if not dry_run:
            self.stdout.write(self.style.SUCCESS(f"Reporte: {report_path}"))
        else:
            self.stdout.write(self.style.WARNING("Dry-run: no se escribió reporte ni se crearon usuarios/equipos."))